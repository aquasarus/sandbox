import os
import argparse
import pyexcel as pe
import openpyxl as op
from openpyxl.styles import Alignment


ETFS = [
    'VBK',
    'VGT',
    'SOXX',
    'FDN',
    'IBB',
    'FBT',
    'IVV',
    'VOO',
    'QCLN',
    'ICLN',
    'BOTZ',
    'GAMR'
]


def get_existing(sheet):
    transactions = set()
    for i in range(2, sheet.max_row + 1):
        transactions.add(sheet.cell(row=i, column=1).value)
    return transactions


def clear_data(sheet):
    for i in range(2, sheet.max_row + 1):
        for j in range(1, 11):
            sheet.cell(row=i, column=j).value = None


def force_alignment(sheet):
    for i in range(1, sheet.max_row + 1):
        for j in range(1, sheet.max_column + 1):
            sheet.cell(row=i, column=j).alignment = Alignment(horizontal='right')


def process_transactions(reset, delete):
    # read transaction data
    filename = '/Users/jasonchan/Downloads/transactions.csv'
    sheet = pe.get_sheet(file_name=filename)

    data = []
    rows = sheet.get_array()

    # skip first and last lines
    for i in range(1, len(rows) - 1):
        row = rows[i]

        if row[4] in ETFS:
            continue # skip ETFs
        elif row[7] < 0 and row[3] < 1:
            continue # skip dividend reinvestments

        data.append({
            'DATE': row[0],
            'ID': row[1],
            'QUANTITY': row[3],
            'TICKER': row[4],
            'PRICE': row[5],
            'COMMISSION': row[6],
            'AMOUNT': row[7]
        })

    # remove transactions file
    if delete:
        os.remove(filename)

    # load tracker excel sheet
    # filename = 'C:/Users/Jason/OneDrive/Documents/Finance/Allowance Tracker.xlsx'
    filename = '/Users/jasonchan/Documents/Allowance Tracker.xlsx'

    wb = op.load_workbook(filename)
    sheet = wb['Transactions']

    if reset:
        row = 2
        clear_data(sheet)
    else:
        row = sheet.max_row + 1 # NOTE: this only works if there is existing data
        existing = get_existing(sheet)

    for record in data:
        transaction_id = int(record['ID'])

        if not reset and transaction_id in existing:
            continue # skip

        sheet.cell(row=row, column=1).value = transaction_id

        date_cell = sheet.cell(row=row, column=2)
        date_cell.style = 'Default'
        date_cell.value = record['DATE']

        amount_cell = sheet.cell(row=row, column=3)
        amount_cell.style = 'Custom Currency'
        amount_cell.value = float(record['AMOUNT'])

        trade_cell = sheet.cell(row=row, column=4)
        trade_cell.style = 'Default'
        if record['AMOUNT'] > 0:
            trade_cell.value = 'SELL'
        else:
            trade_cell.value = 'BUY'

        ticker_cell = sheet.cell(row=row, column=5)
        ticker_cell.style = 'Default'
        ticker_cell.value = record['TICKER']

        quantity_cell = sheet.cell(row=row, column=6)
        quantity_cell.style = 'Default'
        quantity_cell.value = float(record['QUANTITY'])

        price_cell = sheet.cell(row=row, column=7)
        price_cell.style = 'Custom Currency'
        price_cell.value = float(record['PRICE'])

        comm_cell = sheet.cell(row=row, column=8)
        comm_cell.style = 'Custom Currency'
        if record['COMMISSION']:
            comm_cell.value = float(record['COMMISSION'])
        else:
            comm_cell.value = None

        # clear gains
        sheet.cell(row=row, column=9).value = None
        sheet.cell(row=row, column=9).style = 'Custom Currency'
        sheet.cell(row=row, column=10).value = None
        sheet.cell(row=row, column=10).style = 'Custom Percent'

        row += 1 # increment

    force_alignment(sheet)
    wb.save(filename)

if __name__ == "__main__":
    # parse arguments
    parser = argparse.ArgumentParser()
    parser.add_argument('-d', '--delete-transactions', action='store_true', default=False)
    parser.add_argument('-r', '--reset', action='store_true', default=False)

    args = parser.parse_args()
    delete = args.delete_transactions
    reset = args.reset

    process_transactions(reset, delete)
