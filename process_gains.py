import openpyxl as op
from openpyxl.styles import Font


def add_position(portfolio, ticker, shares, price, amount):
    if ticker not in portfolio:
        portfolio[ticker] = {
            'positions': [],
            'drip': []
        }

    portfolio[ticker]['positions'].append({
        'shares': shares,
        'price': price,
        'cost': amount
    })


def sell_lot(positions, shares):
    position = positions[0]
    shares_to_sell = min(shares, position['shares'])

    cost_basis = (shares_to_sell / position['shares']) * position['cost']

    if shares_to_sell == position['shares']:
        positions.pop(0)
    else:
        position['shares'] -= shares_to_sell
        position['cost'] -= cost_basis

    return shares_to_sell, cost_basis


def reduce_position(portfolio, ticker, shares, amount):
    positions = portfolio[ticker]['positions']
    total_cost = 0

    while shares > 0:
        sold, cost_basis = sell_lot(positions, shares)
        shares -= sold
        total_cost += cost_basis

    gains = amount - total_cost

    # process dividends if closing
    # if len(positions) == 0:
    #     drip_amount = 0
    #     for drip in portfolio[ticker]['drip']:
    #         drip_amount += drip * price

    #     portfolio.pop(ticker)
    #     total_gains + drip_amount

    percent_gains = gains / total_cost
    return gains, percent_gains


def main():
    filename = '/Users/jasonchan/Documents/Allowance Tracker.xlsx'
    wb = op.load_workbook(filename)
    sheet = wb['Transactions']

    portfolio = {}

    for row in sheet.iter_rows(min_row=2, max_col=7, max_row=sheet.max_row):
        amount = abs(float(row[2].value))
        buy = (row[3].value == 'BUY')
        shares = float(row[5].value)
        ticker = row[4].value
        price = float(row[6].value)

        # clear first
        sheet.cell(row=row[0].row, column=9).value = None
        sheet.cell(row=row[0].row, column=10).value = None

        if shares < 1:
            continue # TODO: process drip properly

        if buy:
            print('Buy {} of {} for {}'.format(shares, ticker, price))
            add_position(portfolio, ticker, shares, price, amount)
        else:
            print('Sell {} of {} for {}'.format(shares, ticker, price))
            gains, percent = reduce_position(portfolio, ticker, shares, amount)
            sheet.cell(row=row[0].row, column=9).value = gains
            sheet.cell(row=row[0].row, column=10).value = percent

            if gains > 0:
                sheet.cell(row=row[0].row, column=9).font = Font(bold=True, color='FF00B050')
                sheet.cell(row=row[0].row, column=10).font = Font(bold=True, color='FF00B050')
            else:
                sheet.cell(row=row[0].row, column=9).font = Font(bold=True, color='FFFF0000')
                sheet.cell(row=row[0].row, column=10).font = Font(bold=True, color='FFFF0000')

    wb.save(filename)

if __name__ == "__main__":
    main()
